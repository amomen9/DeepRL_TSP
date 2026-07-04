"""
Helper_excel.py - Excel workbook I/O, formatting, and signature/comparison
                  utilities for experiment results.

Contents (high-level groups)
----------------------------
Constants                     - PROJECT_CONFIG_SHEET, EXCLUDED_*_KEYS,
                                sample-template paths and row-height defaults.
Display-width / autosize      - _excel_cell_display_width, _autosize_excel_columns,
                                _set_raw_excel_column_widths.
HP-text normalization         - _sheet_hp_text, _sheet_hp_text_candidates,
                                _filter_config_for_comparison, _parse_sheet_value,
                                _hp_value_matches, _values_equal,
                                _normalize_for_signature, _value_to_text,
                                _job_signature, _sheet_signature.
Project-config sheet          - _write_project_config_sheet,
                                _read_project_config_sheet,
                                _verify_project_config.
Workbook I/O (results)        - _empty_data_sheets_dir, _save_results_to_excel,
                                _load_results_from_excel, _load_all_excel_curves,
                                _parse_sheet_entry,
                                load_algorithm_workbook, save_algorithm_workbook.
Sheet-style helpers           - _load_template_sheet, _copy_style,
                                _apply_excel_sheet_shell, _format_excel,
                                _write_raw_excel_sheet, _next_setting_sheet_name,
                                _rows_from_result, _build_headers, _build_rows.
Benchmark loader              - _load_benchmark_curve.
"""
import os
import ast
import glob
import json
import random
from copy import copy
from datetime import datetime
from functools import lru_cache
from typing import Any, cast

import numpy as np
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

from .Helper_legend import (
    _fmt,
    _format_legend_label,
    _format_legend_value,
    _resolve_legend_flags,
)
from .Library_env_elements import (
    generate_random_duration_matrix,
    generate_random_potential_uncertainty_matrix,
    generate_random_inclusion_matrix,
    env_matrices_text_and_hash,
)


# ── Project-config metadata sheet (workbook-wide config snapshot) ─────────────

PROJECT_CONFIG_SHEET = "__project_config__"

EXCLUDED_GLOBAL_CONFIG_KEYS = frozenset({
    "MIN_UNUSED_CPU_CORES",
    "benchmark_curve",
    "benchmark_name",
    "plot_smoothing_window",
    "curve_confidence_interval",
    "curve_shaded_area_opacity",
    "curve_plot",
    "animation_plot",
    "use_existing_disk_data",
    "Environment",
    "baseline_model",
    "n_use_trained_model",
    "action_selection_method",
    "trained_model_reseed_seed",
    "n_repetitions",
    "k_order_aggregation_methods",
})

EXCLUDED_ALGO_CONFIG_KEYS = frozenset({
    "legend_parameters",
    # The repetition count never gates result reuse: a workbook is matched on the
    # instance + standards + hyperparameters regardless of how many repetitions it
    # holds. Callers load every repetition stored on disk and top up any shortfall
    # against their target by training only the difference (see
    # Library_bello_baseline.run_bello_baseline).
    "n_repetitions",
})

# Project-config sheet rows describing the instance-defining matrices. The full
# matrices are stored as text (provenance), while ``INSTANCE_MATRIX_HASH_KEY`` is
# the value actually compared when deciding whether on-disk results match the
# running instance. Matrix matching for Excel results is unconditional.
INSTANCE_SCOPE = "instance"
INSTANCE_MATRIX_HASH_KEY = "instance_matrix_hash"
INSTANCE_MATRIX_TEXT_KEY = "instance_matrices_text"


# ── Workbook style constants (sample-template defaults) ───────────────────────

SAMPLE_TEMPLATE_PATH = os.path.join("data sheets", "Sample format", "1399.03.12.xlsx")
SETTING_SHEET_PREFIX = "Setting_"

TITLE_ROW_HEIGHT = 34.25
HEADER_ROW_HEIGHT = 52.5
DATA_ROW_HEIGHT = 23.4
DEFAULT_ROW_HEIGHT = 14.5
DEFAULT_COLUMN_WIDTH = 8.90625
SHEET_ZOOM_SCALE = 85


# ── Display-width / autosize helpers ──────────────────────────────────────────

def _excel_cell_display_width(value: Any) -> int:
    """Return the visible width of a value when exported to Excel."""
    if value is None:
        return 0
    if isinstance(value, (float, np.floating)) and np.isnan(value):
        return 0
    text = str(value)
    if not text:
        return 0
    return max(len(line) for line in text.splitlines())


def _autosize_excel_columns(worksheet, dataframe) -> None:
    """Resize Excel columns so their contents fit the widest cell in each column."""
    for column_index, column_name in enumerate(dataframe.columns, start=1):
        max_width = _excel_cell_display_width(column_name)
        for value in dataframe[column_name].tolist():
            max_width = max(max_width, _excel_cell_display_width(value))
        worksheet.column_dimensions[get_column_letter(column_index)].width = max_width


def _set_raw_excel_column_widths(worksheet, headers, first_row_values) -> None:
    """Set raw-sheet widths from the header and first data row only."""
    if first_row_values is None:
        first_row_values = []
    for column_index, header in enumerate(headers, start=1):
        first_value = first_row_values[column_index - 1] if column_index - 1 < len(first_row_values) else ""
        width = max(_excel_cell_display_width(header), _excel_cell_display_width(first_value)) + 1
        worksheet.column_dimensions[get_column_letter(column_index)].width = width


# ── HP-text normalization / signature helpers ─────────────────────────────────

def _sheet_hp_text(value: Any) -> str:
    """Return a stable string representation for Excel hyperparameter values."""
    if value is None:
        return ""
    if isinstance(value, np.ndarray):
        if value.ndim == 0:
            value = value.item()
        else:
            value = value.tolist()
    if isinstance(value, (list, tuple, set)):
        items = list(value)
        if len(items) == 0:
            return "[]"
        if len(items) == 1:
            return _sheet_hp_text(items[0])
        return "[" + ",".join(_sheet_hp_text(item) for item in items) + "]"
    if isinstance(value, np.bool_):
        value = bool(value)
    if isinstance(value, bool):
        return str(value)
    if isinstance(value, (np.integer, int)):
        return str(int(value))
    if isinstance(value, (np.floating, float)):
        number = float(value)
        if number.is_integer():
            return str(int(number))
        return format(number, ".6g")
    if isinstance(value, str):
        text = value.strip()
        if text.startswith("[") and text.endswith("]"):
            try:
                parsed = ast.literal_eval(text)
                return _sheet_hp_text(parsed)
            except Exception:
                return text.replace(", ", ",")
        return text
    return str(value).strip()


def _sheet_hp_text_candidates(value: Any) -> list[str]:
    """Return the normalized text candidate(s) for a config value."""
    return [_sheet_hp_text(value)]


def _filter_config_for_comparison(config: dict[str, Any] | None, excluded_keys: frozenset[str]) -> dict[str, str]:
    """Return {key: normalized_text} for all non-excluded entries in `config`."""
    if not config:
        return {}
    out: dict[str, str] = {}
    for key, value in config.items():
        if key in excluded_keys:
            continue
        out[str(key)] = _sheet_hp_text(value)
    return out


def _parse_sheet_value(val):
    """Parse a value read from an Excel HP column into a stable string."""
    return _sheet_hp_text(val)


def _normalize_for_signature(value: Any) -> Any:
    """Normalize values so workbook values and config values can be compared reliably."""
    if isinstance(value, np.ndarray):
        return tuple(_normalize_for_signature(v) for v in value.tolist())
    if isinstance(value, (list, tuple)):
        return tuple(_normalize_for_signature(v) for v in value)
    if isinstance(value, set):
        return tuple(sorted(_normalize_for_signature(v) for v in value))
    if isinstance(value, np.bool_):
        value = bool(value)
    elif isinstance(value, (np.integer, np.floating)):
        value = value.item()
    return _value_to_text(value)


def _value_to_text(value: Any) -> str:
    """Convert a value to a compact display string suitable for Excel text cells."""
    if value is None:
        return ""
    if isinstance(value, np.ndarray):
        return str(value.tolist())
    if isinstance(value, (list, tuple)):
        return str(list(value))
    if isinstance(value, set):
        return str(sorted(value))
    if isinstance(value, (np.integer, np.floating)):
        return _fmt(value.item())
    if isinstance(value, bool):
        return "True" if value else "False"
    return str(value)


def _job_signature(job_hyperparams: dict[str, Any]) -> tuple[tuple[str, Any], ...]:
    """Create a hashable signature from a job hyperparameter dictionary."""
    return tuple(sorted((str(key), _normalize_for_signature(value)) for key, value in job_hyperparams.items()))


def _sheet_signature(sheet_hyperparams: dict[str, Any]) -> tuple[tuple[str, Any], ...]:
    """Create a hashable signature from parsed workbook hyperparameters."""
    return tuple(sorted((str(key), _normalize_for_signature(value)) for key, value in sheet_hyperparams.items()))


# ── Project-config sheet read/write/verify ────────────────────────────────────

def _write_project_config_sheet(workbook, global_config: dict[str, Any] | None, algo_config: dict[str, Any] | None) -> None:
    """Write a metadata sheet describing the global/algo config snapshot."""
    worksheet = workbook.create_sheet(title=PROJECT_CONFIG_SHEET)
    worksheet.cell(1, 1, "scope")
    worksheet.cell(1, 2, "key")
    worksheet.cell(1, 3, "value")

    row = 2
    for key, value_text in sorted(_filter_config_for_comparison(global_config, EXCLUDED_GLOBAL_CONFIG_KEYS).items()):
        worksheet.cell(row, 1, "global")
        worksheet.cell(row, 2, key)
        worksheet.cell(row, 3, value_text)
        row += 1
    for key, value_text in sorted(_filter_config_for_comparison(algo_config, EXCLUDED_ALGO_CONFIG_KEYS).items()):
        worksheet.cell(row, 1, "algo")
        worksheet.cell(row, 2, key)
        worksheet.cell(row, 3, value_text)
        row += 1

    # Instance-defining matrices: full text (provenance) plus the compared hash.
    env = (global_config or {}).get("Environment")
    matrices_text, matrices_hash = env_matrices_text_and_hash(env)
    if matrices_hash is not None:
        worksheet.cell(row, 1, INSTANCE_SCOPE)
        worksheet.cell(row, 2, INSTANCE_MATRIX_HASH_KEY)
        worksheet.cell(row, 3, matrices_hash)
        row += 1
        worksheet.cell(row, 1, INSTANCE_SCOPE)
        worksheet.cell(row, 2, INSTANCE_MATRIX_TEXT_KEY)
        worksheet.cell(row, 3, matrices_text)
        row += 1

    worksheet.column_dimensions[get_column_letter(1)].width = 10
    worksheet.column_dimensions[get_column_letter(2)].width = 28
    worksheet.column_dimensions[get_column_letter(3)].width = 60
    worksheet.sheet_state = "hidden"


def _read_project_config_sheet(
    filepath: str,
) -> tuple[dict[str, str] | None, dict[str, str] | None, dict[str, str]]:
    """Read the project-config metadata sheet from a workbook.

    Returns ``(global_dict, algo_dict, instance_dict)``. ``global_dict`` and
    ``algo_dict`` are None when the sheet is absent or the workbook can't be
    opened; ``instance_dict`` is always a dict (empty when no instance rows are
    present, e.g. legacy workbooks).
    """
    try:
        workbook = load_workbook(filepath, read_only=True, data_only=True)
    except Exception:
        return None, None, {}
    try:
        if PROJECT_CONFIG_SHEET not in workbook.sheetnames:
            return None, None, {}
        worksheet = workbook[PROJECT_CONFIG_SHEET]
        global_dict: dict[str, str] = {}
        algo_dict: dict[str, str] = {}
        instance_dict: dict[str, str] = {}
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            if row is None or len(row) < 3:
                continue
            scope, key, value = row[0], row[1], row[2]
            if scope is None or key is None:
                continue
            key_text = str(key)
            value_text = "" if value is None else str(value)
            if scope == "global":
                global_dict[key_text] = value_text
            elif scope == "algo":
                algo_dict[key_text] = value_text
            elif scope == INSTANCE_SCOPE:
                instance_dict[key_text] = value_text
        return global_dict, algo_dict, instance_dict
    finally:
        workbook.close()


def _verify_project_config(
    filepath: str,
    global_config: dict[str, Any] | None,
    algo_config: dict[str, Any] | None,
) -> dict[str, tuple[Any, Any]]:
    """Compare on-disk project config metadata to the running project config.

    Returns a dict {key: (disk_value, running_value)} of mismatches. An empty
    dict means everything matched. The special key ``__missing_metadata__``
    indicates the metadata sheet is absent (legacy workbook).

    Rules:
    - Keys in ``EXCLUDED_GLOBAL_CONFIG_KEYS`` / ``EXCLUDED_ALGO_CONFIG_KEYS`` are skipped.
    - For ``global:n_timesteps`` the disk value must be >= the running value
      (longer training on disk is acceptable; shorter is not).
    - The instance-defining matrices (duration / inclusion / potential-uncertainty)
      must match: the workbook's recorded ``instance_matrix_hash`` must equal the
      running environment's. A workbook produced on a different instance (or a
      legacy workbook with no recorded matrices) is rejected. This check is
      skipped only when the running config carries no Environment to hash.
    - All other keys must match by normalized text.
    """
    disk_global, disk_algo, disk_instance = _read_project_config_sheet(filepath)
    if disk_global is None and disk_algo is None:
        return {"__missing_metadata__": ("(not present)", "(required)")}

    running_global = _filter_config_for_comparison(global_config, EXCLUDED_GLOBAL_CONFIG_KEYS)
    running_algo = _filter_config_for_comparison(algo_config, EXCLUDED_ALGO_CONFIG_KEYS)

    mismatches: dict[str, tuple[Any, Any]] = {}

    # Instance-matrix gate: compare the running environment's matrix hash against
    # the one recorded in the workbook. Only enforced when we can hash the
    # running environment (otherwise there is nothing to compare against).
    _running_env = (global_config or {}).get("Environment")
    _, running_matrix_hash = env_matrices_text_and_hash(_running_env)
    if running_matrix_hash is not None:
        disk_matrix_hash = (disk_instance or {}).get(INSTANCE_MATRIX_HASH_KEY)
        if disk_matrix_hash != running_matrix_hash:
            mismatches[f"{INSTANCE_SCOPE}:{INSTANCE_MATRIX_HASH_KEY}"] = (
                disk_matrix_hash if disk_matrix_hash is not None else "(not present)",
                running_matrix_hash,
            )

    for key in sorted(set(running_global) | set(disk_global or {})):
        # Skip excluded keys on the disk side too: a stale key written by an older
        # run (before it was excluded) must not resurface as a phantom mismatch.
        if key in EXCLUDED_GLOBAL_CONFIG_KEYS:
            continue
        disk_val = (disk_global or {}).get(key)
        run_val = running_global.get(key)
        if key == "n_timesteps":
            try:
                disk_n = int(float(disk_val)) if disk_val not in (None, "") else None
                run_n = int(float(run_val)) if run_val not in (None, "") else None
            except (TypeError, ValueError):
                disk_n = None
                run_n = None
            if disk_n is None or run_n is None or disk_n < run_n:
                mismatches[f"global:{key}"] = (disk_val, run_val)
            continue
        if disk_val != run_val:
            mismatches[f"global:{key}"] = (disk_val, run_val)

    for key in sorted(set(running_algo) | set(disk_algo or {})):
        if key in EXCLUDED_ALGO_CONFIG_KEYS:
            continue
        disk_val = (disk_algo or {}).get(key)
        run_val = running_algo.get(key)
        if disk_val != run_val:
            mismatches[f"algo:{key}"] = (disk_val, run_val)

    return mismatches


# ── Workbook I/O for results sheets ───────────────────────────────────────────

def _load_results_from_excel(
    filepath,
    algo_config: dict[str, Any] | None,
    *,
    global_config: dict[str, Any] | None = None,
    formatted_sheets: bool = False,
):
    """Load results from an Excel file, validating each sheet's hyperparameters
    against the current algo config.

    Workbook layout assumption:
    - Row 1 is the title / merged heading row.
    - Row 2 contains the actual column headers.
    - Data starts on row 3.

    Returns a tuple (results, mismatches):
        results: list of dicts [{"learning_curve", "learning_curve_std", "timesteps", "curve_label"}, ...]
        mismatches: dict {param_name: (sheet_value, config_value)} for the first mismatch
                    encountered per parameter across all skipped sheets.
    Sheets that don't match are skipped.

    When ``global_config`` is provided the workbook's project-config metadata
    sheet must match the running global/algo config (with the documented
    exclusions and the ``n_timesteps`` >= rule). Mismatches reject the entire
    workbook.
    """
    algo_config = algo_config or {}
    if global_config is not None:
        project_mismatches = _verify_project_config(filepath, global_config, algo_config)
        if project_mismatches:
            return [], project_mismatches

    header_row = 1 if formatted_sheets else 0
    try:
        sheets = pd.read_excel(filepath, sheet_name=None, engine="openpyxl", header=header_row)
    except Exception as exc:
        raise ValueError(f"Failed to read Excel file '{filepath}': {exc}") from exc
    sheets.pop(PROJECT_CONFIG_SHEET, None)

    def _extract_matching_results(sheet_map):
        basename = os.path.basename(filepath)
        algo_prefix = os.path.splitext(basename)[0].upper()
        legend: dict[str, tuple[str, bool]] = _resolve_legend_flags(algo_config or {}, warn_on_suppression=False)
        skip_keys = {"legend_parameters", "nn_include_hp_in_legend", "nn_include_lr_in_legend"}

        extracted_results = []
        extracted_mismatches = {}
        required_columns = {"timestep", "learning_curve_mean", "learning_curve_std"}

        for sheet_name in sorted(sheet_map.keys()):
            df = sheet_map[sheet_name]
            if not required_columns.issubset(set(df.columns)):
                continue

            sheet_matched = True
            if algo_config:
                for cfg_key, cfg_val in algo_config.items():
                    if cfg_key in skip_keys:
                        continue
                    if cfg_key not in df.columns:
                        continue

                    sheet_val = df[cfg_key].iloc[0]
                    sheet_val_text = _sheet_hp_text(_parse_sheet_value(sheet_val))
                    cfg_val_candidates = _sheet_hp_text_candidates(cfg_val)

                    if sheet_val_text not in cfg_val_candidates:
                        sheet_matched = False
                        if cfg_key not in extracted_mismatches:
                            extracted_mismatches[cfg_key] = (sheet_val_text, cfg_val)

            if not sheet_matched:
                continue

            raw_returns = None
            try:
                timesteps = df["timestep"].values.astype(np.int32)
                learning_curve = df["learning_curve_mean"].values.astype(np.float32)
                learning_curve_std = df["learning_curve_std"].values.astype(np.float32)

                # If present, extract per-repetition learning curves.
                rep_cols = []
                for c in df.columns:
                    c_text = str(c)
                    if c_text.startswith("rep_"):
                        suffix = c_text[len("rep_"):]
                        if suffix.isdigit():
                            rep_cols.append((int(suffix), c_text))
                if rep_cols:
                    rep_cols.sort(key=lambda x: x[0])
                    rep_values = [
                        df[cname].values.astype(np.float32)
                        for _, cname in rep_cols
                    ]
                    raw_returns = np.stack(rep_values, axis=0)  # (n_repetitions, n_eval_points)
            except Exception:
                continue

            label_parts = [algo_prefix]
            for legend_key, (legend_label, show) in legend.items():
                if not show:
                    continue
                if legend_key in df.columns:
                    val = _parse_sheet_value(df[legend_key].iloc[0])
                elif legend_key in algo_config:
                    val = algo_config[legend_key]
                else:
                    continue

                legend_label = _format_legend_label(legend_label)
                if isinstance(val, (bool, np.bool_)):
                    label_parts.append(f"{legend_label}{'✓' if val else '✗'}")
                else:
                    label_parts.append(f"{legend_label}{_format_legend_value(val)}")

            extracted_results.append({
                "learning_curve": learning_curve,
                "learning_curve_std": learning_curve_std,
                "timesteps": timesteps,
                "raw_returns": raw_returns,
                "curve_label": ", ".join(label_parts),
            })

        return extracted_results, extracted_mismatches

    results, mismatches = _extract_matching_results(sheets)

    # Fallback for any older workbook layout that uses the opposite header row.
    if not results:
        try:
            fallback_header = 0 if formatted_sheets else 1
            fallback_sheets = pd.read_excel(filepath, sheet_name=None, engine="openpyxl", header=fallback_header)
        except Exception:
            return results, mismatches
        fallback_sheets.pop(PROJECT_CONFIG_SHEET, None)
        fallback_results, fallback_mismatches = _extract_matching_results(fallback_sheets)
        if fallback_results:
            return fallback_results, fallback_mismatches

    return results, mismatches


def _load_all_excel_curves(
    data_sheets_dir,
    algo_configs=None,
    *,
    global_config: dict[str, Any] | None = None,
    formatted_sheets: bool = False,
):
    """Load all .xlsx files from data_sheets_dir (non-recursive).

    algo_configs maps each workbook stem to its corresponding config dict.
    Sheets are filtered by hyperparameter matching and labels are built by
    _load_results_from_excel.

    When ``global_config`` is provided, each workbook's project-config metadata
    sheet must match the running global/algo config; files that don't match
    are silently skipped with a message.

    Returns a list of dicts: [{curve_label, learning_curve, learning_curve_std, timesteps, source_file}, ...]
    """
    all_curves = []
    pattern = os.path.join(data_sheets_dir, "*.xlsx")
    for filepath in sorted(glob.glob(pattern)):
        basename = os.path.basename(filepath)
        algo_prefix = os.path.splitext(basename)[0].upper()
        algo_config = (algo_configs or {}).get(algo_prefix)
        try:
            results, mismatches = _load_results_from_excel(
                filepath,
                algo_config,
                global_config=global_config,
                formatted_sheets=formatted_sheets,
            )
        except Exception:
            continue
        if global_config is not None and not results and mismatches:
            parts = []
            for param, value_pair in mismatches.items():
                sheet_val, cfg_val = value_pair
                parts.append(f"{param} (Disk data: {sheet_val}, Config: {cfg_val})")
            print(
                f"[{algo_prefix}] Skipping '{basename}' - project-config mismatch: "
                + "; ".join(parts)
            )
            continue
        for entry in results:
            entry["source_file"] = basename
            all_curves.append(entry)
    return all_curves


# ── Sheet styling / formatting helpers ────────────────────────────────────────

@lru_cache(maxsize=1)
def _load_template_sheet():
    """Load the sample formatting sheet once and reuse it for all exports."""
    if not os.path.isfile(SAMPLE_TEMPLATE_PATH):
        return None
    workbook = load_workbook(SAMPLE_TEMPLATE_PATH)
    return workbook[workbook.sheetnames[0]]


def _copy_style(source_cell, target_cell) -> None:
    """Copy all visible style attributes from one cell to another."""
    target_cell._style = copy(source_cell._style)
    if source_cell.has_style:
        target_cell.font = copy(source_cell.font)
        target_cell.fill = copy(source_cell.fill)
        target_cell.border = copy(source_cell.border)
        target_cell.alignment = copy(source_cell.alignment)
        target_cell.protection = copy(source_cell.protection)
        target_cell.number_format = source_cell.number_format


def _apply_excel_sheet_shell(worksheet) -> None:
    """Apply sheet-level settings that match the sample workbook."""
    worksheet.sheet_view.zoomScale = SHEET_ZOOM_SCALE
    worksheet.sheet_format.defaultRowHeight = DEFAULT_ROW_HEIGHT
    worksheet.sheet_format.defaultColWidth = DEFAULT_COLUMN_WIDTH
    worksheet.freeze_panes = None
    worksheet.auto_filter.ref = None
    worksheet.row_dimensions[1].height = TITLE_ROW_HEIGHT
    worksheet.row_dimensions[2].height = HEADER_ROW_HEIGHT
    for row_index in range(3, worksheet.max_row + 1):
        worksheet.row_dimensions[row_index].height = DATA_ROW_HEIGHT


def _format_excel(
    worksheet,
    *,
    algo_name: str,
    hyperparams: dict[str, Any],
    headers: list[str],
    rows: list[list[Any]],
) -> None:
    """Format a worksheet to match the attached sample workbook style."""
    template_sheet = _load_template_sheet()
    n_columns = max(1, len(headers))

    _apply_excel_sheet_shell(worksheet)

    title_text = f"{str(algo_name).upper()}: " + ", ".join(
        f"{key}={_value_to_text(value)}" for key, value in hyperparams.items()
    )
    title_end = get_column_letter(n_columns)
    worksheet.merge_cells(f"A1:{title_end}1")
    worksheet["A1"] = title_text

    if template_sheet is not None:
        _copy_style(template_sheet["A1"], worksheet["A1"])
    else:
        worksheet["A1"].alignment = worksheet["A1"].alignment.copy(horizontal="center", vertical="center")
        worksheet["A1"].font = worksheet["A1"].font.copy(bold=True)

    for column_index, header in enumerate(headers, start=1):
        cell = worksheet.cell(2, column_index, header)
        if template_sheet is not None:
            template_col = min(column_index, template_sheet.max_column)
            _copy_style(template_sheet.cell(2, template_col), cell)
        else:
            cell.alignment = cell.alignment.copy(horizontal="center", vertical="center", wrap_text=True)
            cell.font = cell.font.copy(bold=True)

    for row_index, row_values in enumerate(rows, start=3):
        worksheet.row_dimensions[row_index].height = DATA_ROW_HEIGHT
        for column_index, value in enumerate(row_values, start=1):
            cell = worksheet.cell(row_index, column_index, value)
            if template_sheet is not None:
                template_col = min(column_index, template_sheet.max_column)
                _copy_style(template_sheet.cell(3, template_col), cell)
            else:
                cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
                cell.font = cell.font.copy(bold=True)

    if rows:
        first_data_row = rows[0]
    else:
        first_data_row = ["" for _ in headers]

    for column_index, header in enumerate(headers, start=1):
        first_value = first_data_row[column_index - 1] if column_index - 1 < len(first_data_row) else ""
        width = max(_excel_cell_display_width(header), _excel_cell_display_width(first_value)) + 1
        worksheet.column_dimensions[get_column_letter(column_index)].width = width

    worksheet.sheet_view.tabSelected = True


def _write_raw_excel_sheet(worksheet, headers: list[str], rows: list[list[Any]]) -> None:
    """Write a worksheet without merged cells or style formatting."""
    for column_index, header in enumerate(headers, start=1):
        worksheet.cell(1, column_index, header)

    for row_index, row_values in enumerate(rows, start=2):
        for column_index, value in enumerate(row_values, start=1):
            worksheet.cell(row_index, column_index, value)

    first_data_row = rows[0] if rows else []
    _set_raw_excel_column_widths(worksheet, headers, first_data_row)


def _next_setting_sheet_name(workbook) -> str:
    """Return the next available Setting_### sheet name."""
    highest_index = 0
    for sheet_name in workbook.sheetnames:
        if not sheet_name.startswith(SETTING_SHEET_PREFIX):
            continue
        suffix = sheet_name[len(SETTING_SHEET_PREFIX):]
        try:
            highest_index = max(highest_index, int(suffix))
        except ValueError:
            continue
    return f"{SETTING_SHEET_PREFIX}{highest_index + 1:03d}"


def _rows_from_result(result: tuple[np.ndarray, np.ndarray, np.ndarray]) -> list[list[Any]]:
    """Convert (learning_curve_mean, learning_curve_std, timesteps) into (t, mean, std) rows."""
    learning_curve, learning_curve_std, timesteps = result
    rows: list[list[Any]] = []
    for timestep, mean_value, std_value in zip(timesteps, learning_curve, learning_curve_std):
        rows.append([timestep, mean_value, std_value])
    return rows


def _build_headers(job_hyperparams: dict[str, Any], *, n_repetitions: int = 0) -> list[str]:
    """Build the Excel column headers for a saved setting."""
    rep_headers = [f"rep_{i + 1}" for i in range(int(n_repetitions))]
    return [
        "timestep",
        "learning_curve_mean",
        "learning_curve_std",
        *rep_headers,
        *[str(key) for key in job_hyperparams.keys()],
        "curve_label",
    ]


def _build_rows(
    job_hyperparams: dict[str, Any],
    curve_label: str,
    result: tuple[np.ndarray, np.ndarray, np.ndarray] | tuple[np.ndarray, np.ndarray, np.ndarray, np.ndarray],
) -> list[list[Any]]:
    """Build the Excel rows for a single setting sheet.

    - If raw per-repetition learning curves are provided (4-tuple), writes rep_1..rep_n columns.
    - Hyperparameter columns and curve_label are written only on the first timestep row.
    """
    learning_curve, learning_curve_std, timesteps = result[0], result[1], result[2]
    raw_returns = result[3] if len(result) == 4 else None

    base_rows = _rows_from_result((learning_curve, learning_curve_std, timesteps))
    hyperparam_values = [_value_to_text(value) for value in job_hyperparams.values()]
    n_hps = len(hyperparam_values)

    if raw_returns is None:
        # Keep column alignment even when raw per-rep curves are not available.
        # rep_* columns are still part of the header, so we must emit exactly
        # n_repetitions placeholder values.
        n_repetitions = int(job_hyperparams.get("n_repetitions", 0) or 0)
        raw_returns_arr = None
    else:
        n_repetitions = int(np.asarray(raw_returns).shape[0])
        raw_returns_arr = np.asarray(raw_returns, dtype=np.float32)

    rows: list[list[Any]] = []
    for row_idx, (timestep, mean_value, std_value) in enumerate(base_rows):
        if raw_returns_arr is None:
            rep_values = [None] * n_repetitions
        else:
            rep_values = [raw_returns_arr[rep_idx, row_idx] for rep_idx in range(n_repetitions)]

        if row_idx == 0:
            hp_values_row = hyperparam_values
            curve_label_row = curve_label
        else:
            hp_values_row = [None] * n_hps
            curve_label_row = None

        rows.append([timestep, mean_value, std_value, *rep_values, *hp_values_row, curve_label_row])

    return rows


def _parse_sheet_entry(worksheet, *, formatted_sheets: bool = False) -> dict[str, Any] | None:
    """Parse one worksheet into the in-memory result format."""
    header_row = 2 if formatted_sheets else 1
    data_start_row = 3 if formatted_sheets else 2

    headers: list[str] = []
    header_to_column: dict[str, int] = {}

    for column_index in range(1, worksheet.max_column + 1):
        header_value = worksheet.cell(header_row, column_index).value
        if header_value is None:
            continue
        header_text = str(header_value)
        if header_text in header_to_column:
            continue
        header_to_column[header_text] = column_index
        headers.append(header_text)

    required = {"timestep", "learning_curve_mean", "learning_curve_std"}
    if not required.issubset(set(headers)):
        return None

    column_data: dict[str, list[Any]] = {header: [] for header in headers}
    for row_index in range(data_start_row, worksheet.max_row + 1):
        non_empty = False
        for header in headers:
            column_index = header_to_column[header]
            value = worksheet.cell(row_index, column_index).value
            if value is not None:
                non_empty = True
            column_data[header].append(value)
        if not non_empty:
            continue

    try:
        timesteps = np.asarray(column_data["timestep"], dtype=np.int32)
        learning_curve = np.asarray(column_data["learning_curve_mean"], dtype=np.float32)
        learning_curve_std = np.asarray(column_data["learning_curve_std"], dtype=np.float32)
    except Exception:
        return None

    # Extract per-repetition learning curves (rep_1..rep_n) when present.
    # These must not participate in hyperparam signatures/matching.
    rep_headers: list[str] = []
    for header in headers:
        if (
            isinstance(header, str)
            and header.startswith("rep_")
            and header[len("rep_"):].isdigit()
        ):
            rep_headers.append(header)
    rep_headers.sort(key=lambda h: int(h[len("rep_"):]))
    raw_returns: np.ndarray | None = None
    if rep_headers:
        try:
            rep_arrays: list[np.ndarray] = []
            for h in rep_headers:
                values = column_data[h]
                rep_values = [
                    float(v) if v not in (None, "") else np.nan
                    for v in values
                ]
                rep_arrays.append(np.asarray(rep_values, dtype=np.float32))
            raw_returns = np.stack(rep_arrays, axis=0)  # (n_repetitions, n_points)
            if not np.any(np.isfinite(raw_returns)):
                raw_returns = None
        except Exception:
            raw_returns = None

    hyperparams: dict[str, Any] = {}
    for header in headers:
        if (
            isinstance(header, str)
            and header.startswith("rep_")
            and header[len("rep_"):].isdigit()
        ):
            continue  # rep_* must not participate in hyperparam signatures/matching

        if header in required or header == "curve_label":
            continue
        values = column_data[header]
        first_value = values[0] if values else None
        if first_value is not None:
            hyperparams[header] = _value_to_text(first_value)

    curve_values = column_data.get("curve_label", [])
    curve_label = next((str(value) for value in curve_values if value not in (None, "")), "")

    return {
        "learning_curve": learning_curve,
        "learning_curve_std": learning_curve_std,
        "timesteps": timesteps,
        "curve_label": curve_label,
        "hyperparams": hyperparams,
        "raw_returns": raw_returns,
    }


def load_algorithm_workbook(
    filepath: str,
    *,
    setting_jobs: list[dict[str, Any]] | None = None,
    formatted_sheets: bool = False,
) -> tuple[list[dict[str, Any] | None], dict[str, tuple[Any, Any]]]:
    """Load matching sheets from a workbook."""
    try:
        workbook = load_workbook(filepath)
    except Exception as exc:
        raise ValueError(f"Failed to read Excel file '{filepath}': {exc}") from exc

    parsed_entries: list[dict[str, Any] | None] = []
    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        entry = _parse_sheet_entry(worksheet, formatted_sheets=formatted_sheets)
        if entry is None:
            entry = _parse_sheet_entry(worksheet, formatted_sheets=not formatted_sheets)
        if entry is not None:
            parsed_entries.append(entry)

    if setting_jobs is None:
        return parsed_entries, {}

    job_map: dict[tuple[tuple[str, Any], ...], int] = {}
    for index, job in enumerate(setting_jobs):
        job_map[_job_signature(job["hyperparams"])] = index

    aligned_entries: list[dict[str, Any] | None] = [None] * len(setting_jobs)
    for entry in parsed_entries:
        if entry is None:
            continue
        signature = _sheet_signature(entry["hyperparams"])
        job_index = job_map.get(signature)
        if job_index is not None and aligned_entries[job_index] is None:
            aligned_entries[job_index] = entry

    return aligned_entries, {}


def save_algorithm_workbook(
    dir_path: str,
    base_filename: str,
    algo_name: str,
    setting_jobs: list[dict[str, Any]],
    setting_results: list[tuple[np.ndarray, np.ndarray, np.ndarray] | None],
    *,
    global_config: dict[str, Any] | None = None,
    algo_config: dict[str, Any] | None = None,
    format_sheets: bool = False,
) -> str:
    """Save or append settings into an algorithm workbook.

    When ``global_config`` / ``algo_config`` are supplied, a hidden metadata
    sheet (``__project_config__``) is written so future loads can verify the
    workbook was produced with matching settings. If an existing workbook on
    disk has incompatible metadata, its previous entries are discarded and the
    file is overwritten with the current run's data.
    """
    os.makedirs(dir_path, exist_ok=True)
    filepath = os.path.join(dir_path, f"{base_filename}.xlsx")

    existing_entries: list[dict[str, Any] | None] = []
    if os.path.isfile(filepath):
        config_ok = True
        if global_config is not None or algo_config is not None:
            project_mismatches = _verify_project_config(
                filepath,
                global_config or {},
                algo_config or {},
            )
            if project_mismatches:
                config_ok = False
                parts = []
                for param, value_pair in project_mismatches.items():
                    sheet_val, cfg_val = value_pair
                    parts.append(f"{param} (Disk data: {sheet_val}, Config: {cfg_val})")
                print(
                    f"[{algo_name}] Existing workbook config differs from running project; "
                    f"discarding previous data. Reason(s): " + "; ".join(parts)
                )
        if config_ok:
            try:
                existing_entries, _ = load_algorithm_workbook(filepath, formatted_sheets=format_sheets)
            except Exception:
                existing_entries = []

    existing_signatures: set[tuple[tuple[str, Any], ...]] = set()
    for entry in existing_entries:
        if entry is None:
            continue
        existing_signatures.add(_sheet_signature(entry["hyperparams"]))

    all_entries = [entry for entry in existing_entries if entry is not None]
    added_count = 0
    for job, result in zip(setting_jobs, setting_results):
        if result is None:
            continue

        job_signature = _job_signature(job["hyperparams"])
        if job_signature in existing_signatures:
            continue

        raw_returns = None
        if isinstance(result, tuple) and len(result) == 4:
            result_4 = cast(tuple[np.ndarray, np.ndarray, np.ndarray, np.ndarray], result)
            learning_curve, learning_curve_std, timesteps, raw_returns = result_4
        else:
            learning_curve, learning_curve_std, timesteps = result

        all_entries.append({
            "learning_curve": learning_curve,
            "learning_curve_std": learning_curve_std,
            "timesteps": timesteps,
            "raw_returns": raw_returns,
            "curve_label": job["curve_label"],
            "hyperparams": job["hyperparams"],
        })
        existing_signatures.add(job_signature)
        added_count += 1

    workbook = Workbook()
    default_sheet = workbook.active
    if default_sheet is not None:
        workbook.remove(default_sheet)

    if global_config is not None or algo_config is not None:
        _write_project_config_sheet(workbook, global_config or {}, algo_config or {})

    if not all_entries:
        worksheet = workbook.create_sheet(title=_next_setting_sheet_name(workbook))
        if format_sheets:
            _format_excel(
                worksheet,
                algo_name=algo_name,
                hyperparams={},
                headers=["timestep", "learning_curve_mean", "learning_curve_std", "curve_label"],
                rows=[],
            )
        else:
            _write_raw_excel_sheet(
                worksheet,
                ["timestep", "learning_curve_mean", "learning_curve_std", "curve_label"],
                [],
            )
    else:
        for index, entry in enumerate(all_entries, start=1):
            worksheet = workbook.create_sheet(title=f"Setting_{index:03d}")

            try:
                n_reps = int(float(entry["hyperparams"].get("n_repetitions", 0)))
            except Exception:
                n_reps = 0

            headers = _build_headers(entry["hyperparams"], n_repetitions=n_reps)

            raw_returns = entry.get("raw_returns")
            if raw_returns is not None:
                result_tuple = (
                    entry["learning_curve"],
                    entry["learning_curve_std"],
                    entry["timesteps"],
                    raw_returns,
                )
            else:
                result_tuple = (
                    entry["learning_curve"],
                    entry["learning_curve_std"],
                    entry["timesteps"],
                )

            rows = _build_rows(entry["hyperparams"], entry["curve_label"], result_tuple)
            if format_sheets:
                _format_excel(
                    worksheet,
                    algo_name=algo_name,
                    hyperparams=entry["hyperparams"],
                    headers=headers,
                    rows=rows,
                )
            else:
                _write_raw_excel_sheet(worksheet, headers, rows)

    workbook.save(filepath)
    print(f"Saved {added_count} new setting(s) to {filepath}")
    return filepath


# ── Benchmark CSV loader ──────────────────────────────────────────────────────

def _load_benchmark_curve(
    benchmark_curve,
    project_eval_interval,
    project_n_timesteps,
    benchmark_eval_interval=250,
    episode_return_column="Episode_Return",
):
    benchmark_files = {
        1: os.path.join("Baseline data", "BaselineDataTSP_run1.csv"),
        2: os.path.join("Baseline data", "BaselineDataTSP_run2.csv"),
    }
    if benchmark_curve not in benchmark_files:
        raise ValueError("benchmark_curve must be 1 or 2.")

    data = np.genfromtxt(benchmark_files[benchmark_curve], delimiter=",", names=True)
    if data.dtype.names is None or episode_return_column not in data.dtype.names:
        raise ValueError(
            f"Selected benchmark file does not contain requested column '{episode_return_column}'."
        )

    env_steps = np.atleast_1d(np.asarray(data["env_step"], dtype=np.float32))
    returns = np.atleast_1d(np.asarray(data[episode_return_column], dtype=np.float32))

    if env_steps.size == 0 or returns.size == 0:
        raise ValueError("Selected benchmark file has no usable data.")

    valid_rows = np.isfinite(env_steps) & np.isfinite(returns)
    env_steps = env_steps[valid_rows]
    returns = returns[valid_rows]

    if env_steps.size == 0:
        raise ValueError("Selected benchmark file contains only invalid rows.")

    sort_idx = np.argsort(env_steps)
    env_steps = env_steps[sort_idx]
    returns = returns[sort_idx]

    if project_eval_interval != benchmark_eval_interval and env_steps.size >= 2:
        normalized_steps = np.arange(
            env_steps[0],
            env_steps[-1] + project_eval_interval,
            project_eval_interval,
            dtype=np.float32,
        )
        normalized_steps = normalized_steps[normalized_steps <= env_steps[-1]]
        returns = np.interp(normalized_steps, env_steps, returns).astype(np.float32)
        env_steps = normalized_steps

    in_horizon = env_steps <= float(project_n_timesteps)
    if not np.any(in_horizon):
        print(
            f"[benchmark] No benchmark points fall within project_n_timesteps={project_n_timesteps}; "
            "using the full benchmark curve instead."
        )
        return env_steps.astype(np.int32), returns

    env_steps = env_steps[in_horizon]
    returns = returns[in_horizon]
    return env_steps.astype(np.int32), returns


# ── Sample-data matrix workbooks ──────────────────────────────────────────────
# A "trio" is the three workbooks produced by one sample-data generation run:
#   <prefix>__<timestamp>__<id>.xlsx, with one matrix per sheet (sheet name =
# matrix dimension). The three files share a timestamp and a unique 3-digit id.

SAMPLE_DATA_DIR = os.path.join("data sheets", "sample data")

SAMPLE_DURATION_PREFIX = "duration_matrix"
SAMPLE_UNCERTAINTY_PREFIX = "potential_uncertainty_matrix"
SAMPLE_INCLUSION_PREFIX = "uncertainty_inclusion_matrix"

_SAMPLE_MATRIX_PREFIXES = (
    SAMPLE_DURATION_PREFIX,
    SAMPLE_UNCERTAINTY_PREFIX,
    SAMPLE_INCLUSION_PREFIX,
)


def _existing_sample_ids(directory: str) -> set[str]:
    """Return the set of 3-digit ids already present in ``directory``."""
    ids: set[str] = set()
    if not os.path.isdir(directory):
        return ids
    for name in os.listdir(directory):
        stem, ext = os.path.splitext(name)
        if ext.lower() != ".xlsx":
            continue
        parts = stem.split("__")
        if len(parts) == 3 and parts[-1].isdigit() and len(parts[-1]) == 3:
            ids.add(parts[-1])
    return ids


def _generate_unique_sample_id(directory: str) -> str:
    """Pick a random unused 3-digit id (100-999) for a new trio."""
    used = _existing_sample_ids(directory)
    available = [f"{i:03d}" for i in range(100, 1000) if f"{i:03d}" not in used]
    if not available:
        raise RuntimeError(f"No unused 3-digit id remains in '{directory}'.")
    return random.choice(available)


def _write_sample_matrix_workbook(prefix, timestamp, file_id, sheets, directory):
    """Write one workbook holding a matrix per sheet (sheet name = dimension).

    ``sheets`` maps dimension (int) -> 2D matrix. Sheets are written in
    ascending order of dimension.
    """
    filename = f"{prefix}__{timestamp}__{file_id}.xlsx"
    filepath = os.path.join(directory, filename)
    with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
        for dimension in sorted(sheets.keys()):
            df = pd.DataFrame(sheets[dimension])
            df.to_excel(writer, sheet_name=str(dimension), header=False, index=False)
    return filepath


def generate_sample_data(config, directory: str = SAMPLE_DATA_DIR):
    """Generate one trio of sample-data workbooks from ``config``.

    ``config`` is a dict with a ``"matrices"`` list of group dicts whose keys
    mirror the arguments of the three matrix generators. The three workbooks
    share a single timestamp and a unique 3-digit id.
    """
    os.makedirs(directory, exist_ok=True)

    timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    file_id = _generate_unique_sample_id(directory)

    groups = sorted(config["matrices"], key=lambda group: group["dimensions"])

    duration_sheets = {}
    uncertainty_sheets = {}
    inclusion_sheets = {}

    for group in groups:
        n = group["dimensions"]

        duration_matrix = generate_random_duration_matrix(
            n=n,
            min_durn=group["min_durn"],
            max_durn=group["max_durn"],
            symmetric=group["duration_symmetric"],
            seed=group["seed"],
        )
        potential_uncertainty_matrix = generate_random_potential_uncertainty_matrix(
            duration_matrix=duration_matrix,
            min_uncertainty=group["min_uncertainty"],
            max_uncertainty=group["max_uncertainty"],
            uncertainty_scale=group["uncertainty_scale"],
            uncertainty_symmetric=group["uncertainty_symmetric"],
            seed=group["seed"],
        )
        uncertainty_inclusion_matrix = generate_random_inclusion_matrix(
            n=n,
            n_uncertain_routes=group["n_uncertain_routes"],
            symmetric=group["inclusion_symmetric"],
            seed=group["seed"],
        )

        duration_sheets[n] = duration_matrix
        uncertainty_sheets[n] = potential_uncertainty_matrix
        inclusion_sheets[n] = uncertainty_inclusion_matrix

    duration_path = _write_sample_matrix_workbook(
        SAMPLE_DURATION_PREFIX, timestamp, file_id, duration_sheets, directory
    )
    uncertainty_path = _write_sample_matrix_workbook(
        SAMPLE_UNCERTAINTY_PREFIX, timestamp, file_id, uncertainty_sheets, directory
    )
    inclusion_path = _write_sample_matrix_workbook(
        SAMPLE_INCLUSION_PREFIX, timestamp, file_id, inclusion_sheets, directory
    )

    print(f"Generated sample-data trio (id={file_id}, timestamp={timestamp}):")
    for path in (duration_path, uncertainty_path, inclusion_path):
        print(f"  {path}")

    return file_id, timestamp


def _discover_sample_trios(directory: str) -> dict:
    """Group sample-data workbooks by 3-digit id.

    Returns {file_id: {"timestamp": str, "files": {prefix: filepath}}}.
    """
    trios: dict = {}
    if not os.path.isdir(directory):
        return trios
    for name in os.listdir(directory):
        stem, ext = os.path.splitext(name)
        if ext.lower() != ".xlsx":
            continue
        parts = stem.split("__")
        if len(parts) != 3:
            continue
        prefix, timestamp, file_id = parts
        if prefix not in _SAMPLE_MATRIX_PREFIXES:
            continue
        if not (file_id.isdigit() and len(file_id) == 3):
            continue
        entry = trios.setdefault(file_id, {"timestamp": timestamp, "files": {}})
        entry["files"][prefix] = os.path.join(directory, name)
    return trios


def load_sample_matrices(dimension, file_id=None, directory: str = SAMPLE_DATA_DIR):
    """Load the duration / uncertainty / inclusion matrices for one trio.

    ``file_id`` selects the trio by its 3-digit id; when None the latest
    created (most recent timestamp) complete trio is used. ``dimension``
    selects which sheet (matrix size) to read. Returns a dict with keys
    ``duration_matrix``, ``potential_uncertainty_matrix`` and
    ``uncertainty_inclusion_matrix``.
    """
    not_found_message = (
        "The problem size with sample data id (if any was given) was not found "
        "on the disk. Use Generate_Sample_Data.py file to generate your desired "
        "sample data."
    )

    trios = _discover_sample_trios(directory)
    complete = {
        fid: info
        for fid, info in trios.items()
        if all(prefix in info["files"] for prefix in _SAMPLE_MATRIX_PREFIXES)
    }
    if not complete:
        raise FileNotFoundError(not_found_message)

    if file_id is None:
        def _trio_recency(item):
            _fid, info = item
            mtimes = [os.path.getmtime(p) for p in info["files"].values()]
            return (info["timestamp"], max(mtimes))

        file_id = max(complete.items(), key=_trio_recency)[0]
    else:
        file_id = f"{int(file_id):03d}"
        if file_id not in complete:
            raise FileNotFoundError(not_found_message)

    info = complete[file_id]
    sheet_name = str(dimension)
    matrices = {}
    for prefix in _SAMPLE_MATRIX_PREFIXES:
        filepath = info["files"][prefix]
        try:
            df = pd.read_excel(
                filepath,
                sheet_name=sheet_name,
                header=None,
                engine="openpyxl",
            )
        except ValueError as exc:
            # pandas raises ValueError when the requested dimension sheet is absent.
            raise FileNotFoundError(not_found_message) from exc
        matrices[prefix] = df.values.tolist()
        print(
            f"[sample data] Loaded {prefix} (dimension {dimension}, id {file_id}) "
            f"from '{filepath}'."
        )
    # Trio provenance used to locate/create the cached DP-solution workbook.
    matrices["file_id"] = file_id
    matrices["timestamp"] = info["timestamp"]
    return matrices


# ── Cached DP solutions (per-trio, per-dimension) ─────────────────────────────
# A trio's classic DP solutions (best / worst / expected tours and costs) are
# cached in a sibling workbook named
#   ``duration_matrix_DP-Solution_<timestamp>__<file id>.xlsx``
# holding one sheet per solved dimension, named ``<dimension>-solution``. Each
# sheet stores the attachment format (Best/Worst value+route(s)) plus the
# project's Expected value+route(s). Routes are encoded as a JSON array of
# ``{"city_order": [...], "route_pairs": [[a, b], ...]}`` objects, 1-indexed and
# starting from the depot. Once written, later runs on the same trio+dimension
# read the solution back instead of recomputing the (2**n) DP.

DP_SOLUTION_PREFIX = "duration_matrix_DP-Solution"
DP_SOLUTION_SHEET_SUFFIX = "-solution"


def _dp_solution_sheet_name(dimension) -> str:
    """Return the solution sheet name for a matrix dimension (e.g. "24-solution")."""
    return f"{dimension}{DP_SOLUTION_SHEET_SUFFIX}"


def _dp_solution_new_path(file_id, timestamp, directory: str = SAMPLE_DATA_DIR) -> str:
    """Build the path for a freshly created DP-solution workbook."""
    fid = f"{int(file_id):03d}"
    return os.path.join(directory, f"{DP_SOLUTION_PREFIX}_{timestamp}__{fid}.xlsx")


def _find_dp_solution_file(file_id, directory: str = SAMPLE_DATA_DIR, timestamp=None):
    """Return the DP-solution workbook path for a trio id, or None if absent.

    When ``timestamp`` is given and the exactly-named sibling exists it is
    preferred; otherwise the most recent matching workbook (by filename) is used.
    """
    if file_id is None or not os.path.isdir(directory):
        return None
    fid = f"{int(file_id):03d}"
    if timestamp is not None:
        exact = _dp_solution_new_path(fid, timestamp, directory)
        if os.path.isfile(exact):
            return exact
    pattern = os.path.join(directory, f"{DP_SOLUTION_PREFIX}_*__{fid}.xlsx")
    matches = sorted(glob.glob(pattern))
    return matches[-1] if matches else None


def _tours_to_route_json(tours) -> str:
    """Encode 0-indexed tours as a JSON array of {city_order, route_pairs} objects.

    ``city_order`` and ``route_pairs`` are 1-indexed (display form), so a tour
    ``[0, 2, 1, 3, 0]`` becomes ``{"city_order": [1, 3, 2, 4, 1],
    "route_pairs": [[1, 3], [3, 2], [2, 4], [4, 1]]}``.
    """
    routes = []
    for tour in tours:
        order = [int(city) + 1 for city in tour]
        pairs = [[order[i], order[i + 1]] for i in range(len(order) - 1)]
        routes.append({"city_order": order, "route_pairs": pairs})
    return json.dumps(routes)


def _route_json_to_tours(text) -> list[list[int]]:
    """Decode the route JSON back to 0-indexed tours (list of city lists)."""
    if text is None or str(text).strip() == "":
        return []
    data = json.loads(text)
    tours: list[list[int]] = []
    for entry in data:
        order = entry["city_order"] if isinstance(entry, dict) else entry
        tours.append([int(city) - 1 for city in order])
    return tours


def _sort_dp_solution_sheets(workbook) -> None:
    """Order the solution sheets by ascending dimension (mirrors the trio)."""
    def _key(worksheet):
        name = worksheet.title
        base = (
            name[: -len(DP_SOLUTION_SHEET_SUFFIX)]
            if name.endswith(DP_SOLUTION_SHEET_SUFFIX)
            else name
        )
        try:
            return (0, int(base))
        except ValueError:
            return (1, name)

    workbook._sheets.sort(key=_key)


def load_dp_solution(file_id, dimension, directory: str = SAMPLE_DATA_DIR):
    """Return the cached DP solution for one trio + dimension, or None if absent.

    The returned dict has keys ``best_cost``, ``best_tours``, ``worst_cost``,
    ``worst_tours``, ``expected_cost``, ``expected_tours``. Tours are 0-indexed
    lists. Returns None when no workbook / sheet exists or it cannot be parsed.
    """
    filepath = _find_dp_solution_file(file_id, directory)
    if filepath is None:
        return None
    sheet_name = _dp_solution_sheet_name(dimension)
    try:
        workbook = load_workbook(filepath, read_only=True, data_only=True)
    except Exception:
        return None
    try:
        if sheet_name not in workbook.sheetnames:
            return None
        worksheet = workbook[sheet_name]
        values: dict[str, Any] = {}
        for row in worksheet.iter_rows(values_only=True):
            if not row or row[0] is None:
                continue
            key = str(row[0]).strip()
            values[key] = row[1] if len(row) > 1 else None
    finally:
        workbook.close()

    def _num(value):
        return float(value) if value is not None else None

    try:
        return {
            "best_cost": _num(values.get("Best value")),
            "best_tours": _route_json_to_tours(values.get("Best route(s)")),
            "worst_cost": _num(values.get("Worst value")),
            "worst_tours": _route_json_to_tours(values.get("Worst route(s)")),
            "expected_cost": _num(values.get("Expected value")),
            "expected_tours": _route_json_to_tours(values.get("Expected route(s)")),
        }
    except Exception:
        return None


def save_dp_solution(
    file_id,
    dimension,
    solution: dict[str, Any],
    *,
    timestamp=None,
    directory: str = SAMPLE_DATA_DIR,
) -> str:
    """Write (or overwrite) the cached DP solution sheet for one trio + dimension.

    ``solution`` maps ``best_cost``/``best_tours``/``worst_cost``/``worst_tours``/
    ``expected_cost``/``expected_tours`` (tours are 0-indexed lists). The workbook
    is created on first use for the trio (named with ``timestamp`` when given,
    otherwise the current time); subsequent dimensions add/replace their own
    ``<dimension>-solution`` sheet in the same file.
    """
    os.makedirs(directory, exist_ok=True)
    filepath = _find_dp_solution_file(file_id, directory, timestamp=timestamp)
    if filepath is None:
        effective_timestamp = timestamp or datetime.now().strftime("%Y%m%d-%H%M%S")
        filepath = _dp_solution_new_path(file_id, effective_timestamp, directory)

    if os.path.isfile(filepath):
        workbook = load_workbook(filepath)
    else:
        workbook = Workbook()
        default_sheet = workbook.active
        if default_sheet is not None:
            workbook.remove(default_sheet)

    sheet_name = _dp_solution_sheet_name(dimension)
    if sheet_name in workbook.sheetnames:
        workbook.remove(workbook[sheet_name])
    worksheet = workbook.create_sheet(title=sheet_name)

    rows = [
        ("Best value", solution["best_cost"]),
        ("Best route(s)", _tours_to_route_json(solution["best_tours"])),
        ("Worst value", solution["worst_cost"]),
        ("Worst route(s)", _tours_to_route_json(solution["worst_tours"])),
        ("Expected value", solution["expected_cost"]),
        ("Expected route(s)", _tours_to_route_json(solution["expected_tours"])),
    ]
    for row_index, (label, value) in enumerate(rows, start=1):
        worksheet.cell(row_index, 1, label)
        worksheet.cell(row_index, 2, value)
    worksheet.column_dimensions["A"].width = 16
    worksheet.column_dimensions["B"].width = 80

    _sort_dp_solution_sheets(workbook)
    workbook.save(filepath)
    return filepath
